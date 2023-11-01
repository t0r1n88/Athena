"""
Скрипт для получения данных по школам муниципалитета
"""
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.chart import BarChart, Reference
import numpy as np
import os
from tkinter import messagebox
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
pd.options.mode.chained_assignment = None  # default='warn'


def create_report_mun(folder_data:str,path_end_folder:str):
    """
    Функция для извлечения данных из отчета по школам муниципалитета
    """
    try:
        current_date = datetime.now()
        formatted_date = current_date.strftime("%d_%m_%Y")
        for file in os.listdir(folder_data):
            if (file.endswith('.xlsx') and not file.startswith('~$')):
                name_file = file.split('.xlsx')[0]  # получаем название файла
                wb_temp = openpyxl.load_workbook(f'{folder_data}/{file}',read_only=True)
                lst_sheet = []  # листы для обработки
                for sheet in wb_temp.sheetnames:
                    if ('43' in sheet) or ('44' in sheet):
                        lst_sheet.append(sheet)
                wb_temp.close()
                    # удаляем лишние листы
                # for sheet_name in wb.sheetnames[2:]:
                #     if not ('43' in sheet_name) or not ('44' in sheet_name):
                #         del wb[sheet_name]

                wb=openpyxl.Workbook()

                wb.create_sheet(title='43 Рейтинг активности', index=1)
                wb.create_sheet(title='44 Активные пользователи', index=2)
                del wb['Sheet']
                df_43 = pd.read_excel(f'{folder_data}/{file}',
                                      sheet_name=lst_sheet[0])  # получаем рейтинг событий по муниципалитету
                #
                df_43.iloc[:, 1] = df_43.iloc[:, 1].astype(float)

                df_43.rename(columns={'Unnamed: 0': 'Школа'}, inplace=True)
                sort_column = df_43.columns[1]

                df_43.sort_values(by=sort_column, ascending=False, inplace=True)

                for row in dataframe_to_rows(df_43, index=False, header=True):
                    wb['43 Рейтинг активности'].append(row)
                # сохраняем по ширине колонок
                for column in wb['43 Рейтинг активности'].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    wb['43 Рейтинг активности'].column_dimensions[column_name].width = adjusted_width

                # рисуем график
                chart = BarChart()
                chart.title = "Диаграмма"
                chart.y_axis.title = 'Количество событий'
                chart.x_axis.title = 'Школа'

                data = Reference(wb['43 Рейтинг активности'], min_col=2, min_row=1, max_row=len(df_43) + 1, max_col=2)
                categories = Reference(wb['43 Рейтинг активности'], min_col=1, min_row=2, max_row=len(df_43) + 1)
                # Установите размер диаграммы
                chart.width = 30  # Установите желаемую ширину диаграммы
                chart.height = 20  # Установите желаемую высоту диаграммы

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.y_axis.title = None
                chart.type = "bar"

                wb['43 Рейтинг активности'].add_chart(chart, "D1")

                # Обрабатываем 44 лист
                df_44 = pd.read_excel(f'{folder_data}/{file}',
                                      sheet_name=lst_sheet[1])  # получаем рейтинг активных школ по муниципалитету
                #
                df_44.iloc[:, 1] = df_44.iloc[:, 1].astype(float)

                df_44.rename(columns={'Unnamed: 0': 'Школа'}, inplace=True)
                sort_column = df_44.columns[1]

                df_44.sort_values(by=sort_column, ascending=False, inplace=True)

                for row in dataframe_to_rows(df_44, index=False, header=True):
                    wb['44 Активные пользователи'].append(row)
                # сохраняем по ширине колонок
                for column in wb['44 Активные пользователи'].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    wb['44 Активные пользователи'].column_dimensions[column_name].width = adjusted_width

                # рисуем график
                chart = BarChart()
                chart.title = "Диаграмма"
                chart.y_axis.title = 'Активные пользователи'
                chart.x_axis.title = 'Школа'

                data = Reference(wb['44 Активные пользователи'], min_col=2, min_row=1, max_row=len(df_44) + 1,
                                 max_col=2)
                categories = Reference(wb['44 Активные пользователи'], min_col=1, min_row=2, max_row=len(df_44) + 1)
                # Установите размер диаграммы
                chart.width = 30  # Установите желаемую ширину диаграммы
                chart.height = 20  # Установите желаемую высоту диаграммы

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.y_axis.title = None
                chart.type = "bar"

                wb['44 Активные пользователи'].add_chart(chart, "D1")

                # Сохраняем с названием

                wb.save(f'{path_end_folder}/Отчет {name_file} от {formatted_date}.xlsx')






    except NameError:
        messagebox.showerror('Афина',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Афина',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Афина',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Афина',
                             f'Закройте открытые файлы Excel {e.args}')
    else:
        messagebox.showinfo('Афина',
                            'Данные успешно обработаны.')


if __name__ =='__main__':
    data_folder_main = 'data/Муниципалитеты'
    end_folder = 'data'
    create_report_mun(data_folder_main,end_folder)

    print('Lindy Booth')