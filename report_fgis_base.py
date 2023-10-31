"""
Скрипт для обработки листа 43 и листа 44 отчета по активности во ФГИС
"""
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time
from openpyxl.chart import BarChart, Reference
import numpy as np
from tkinter import messagebox



def create_base_report(data_file:str,path_end_folder:str):
    """
    Функция для генерации в удобном виде значений из листов 43 и 44 отчета по ФГИС Моя Школа
    :param data_file: путь к файлу с данными
    :param path_end_folder:  путь к конечной папке
    :return:
    """
    try:
        wb_temp = openpyxl.load_workbook(data_file, read_only=True)
        lst_sheet = []  # листы для обработки
        for sheet in wb_temp.sheetnames:
            if ('43' in sheet) or ('44' in sheet):
                lst_sheet.append(sheet)
        wb_temp.close()

        wb = openpyxl.Workbook()  # создаем итоговый файл
        wb.create_sheet(title='43 Доля активных школ', index=0)
        wb.create_sheet(title='44 Доля активных польз', index=1)

        del wb['Sheet']

        df_43 = pd.read_excel(data_file, sheet_name=lst_sheet[0])  # получаем рейтинг активных школ по муниципалитетам
        df_43.iloc[:, 1] = df_43.iloc[:, 1].astype(float)

        df_43.rename(columns={'Unnamed: 0': 'Муниципалитет'}, inplace=True)

        df_43.sort_values(by='доля активных школ от зарегистрированных в "Моей школе"', ascending=False, inplace=True)

        for row in dataframe_to_rows(df_43, index=False, header=True):
            wb['43 Доля активных школ'].append(row)
        # сохраняем по ширине колонок
        for column in wb['43 Доля активных школ'].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb['43 Доля активных школ'].column_dimensions[column_name].width = adjusted_width

        chart = BarChart()
        chart.title = "Диаграмма"
        chart.y_axis.title = 'Процент активных школ'
        chart.x_axis.title = 'Муниципалитет'

        data = Reference(wb['43 Доля активных школ'], min_col=2, min_row=1, max_row=len(df_43) + 1, max_col=2)
        categories = Reference(wb['43 Доля активных школ'], min_col=1, min_row=2, max_row=len(df_43) + 1)
        # Установите размер диаграммы
        chart.width = 30  # Установите желаемую ширину диаграммы
        chart.height = 20  # Установите желаемую высоту диаграммы

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.y_axis.title = None
        chart.type = "bar"

        wb['43 Доля активных школ'].add_chart(chart, "D1")

        # 44
        df_44 = pd.read_excel(data_file, sheet_name=lst_sheet[1])  # получаем рейтинг активных школ по муниципалитетам
        df_44.iloc[:, 1] = df_44.iloc[:, 1].astype(float)

        df_44.rename(columns={'Unnamed: 0': 'Муниципалитет'}, inplace=True)

        df_44.sort_values(by='доля активных пользователей в "Моей школе"', ascending=False, inplace=True)

        for row in dataframe_to_rows(df_44, index=False, header=True):
            wb['44 Доля активных польз'].append(row)
        # сохраняем по ширине колонок
        for column in wb['44 Доля активных польз'].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb['44 Доля активных польз'].column_dimensions[column_name].width = adjusted_width

        chart = BarChart()
        chart.title = "Диаграмма"
        chart.y_axis.title = 'Процент активных пользователей'
        chart.x_axis.title = 'Муниципалитет'

        data = Reference(wb['44 Доля активных польз'], min_col=2, min_row=1, max_row=len(df_44) + 1, max_col=2)
        categories = Reference(wb['44 Доля активных польз'], min_col=1, min_row=2, max_row=len(df_44) + 1)
        # Установите размер диаграммы
        chart.width = 30  # Установите желаемую ширину диаграммы
        chart.height = 20  # Установите желаемую высоту диаграммы

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.y_axis.title = None
        chart.type = "bar"

        wb['44 Доля активных польз'].add_chart(chart, "D1")

        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        wb.save(f'{path_end_folder}/Данные от {current_time}.xlsx')
    except NameError:
        messagebox.showerror('Афина',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except KeyError as e:
        messagebox.showerror('Афина',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Афина',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Афина',
                             f'Закройте открытые файлы Excel {e.args}')
    else:
        messagebox.showinfo('Афина',
                            'Данные успешно обработаны.')


if __name__ =='__main__':
    data_file_main = 'data/Аналитика 30_10_2023.xlsx'
    end_folder = 'data'
    create_base_report(data_file_main,end_folder)

    print('Lindy Booth')